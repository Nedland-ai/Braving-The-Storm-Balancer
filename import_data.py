import json
import config
from openpyxl import load_workbook


def parse_resources(cell):
    """Convert 'WOOD=2;STONE=3' -> {'WOOD': 2, 'STONE': 3}"""

    if not cell:
        return {}

    result = {}

    for item in str(cell).split(";"):
        item = item.strip()

        if not item:
            continue

        key, value = item.split("=")
        result[key.strip()] = int(value)

    return result


def parse_list(cell):
    """Convert 'A;B;C' -> ['A', 'B', 'C']"""

    if not cell:
        return []

    return [
        item.strip()
        for item in str(cell).split(";")
        if item.strip()
    ]


def load_projects():
    wb = load_workbook(config.PROJECTS_FILE_PATH)
    ws = wb["projets"]

    projects = []

    for row in ws.iter_rows(min_row=2, values_only=True):

        (
            pid,
            card,
            name,
            workers,
            requirements,
            inputs,
            outputs,
            grants,
            tags
        ) = row

        projects.append({

            "id": pid,
            "card": card,
            "name": name,
            "workers_required": workers or 0,
            "requirements": parse_list(requirements),
            "inputs": parse_resources(inputs),
            "outputs": parse_resources(outputs),
            "grants": parse_list(grants),
            "tags": parse_list(tags)

        })

    with open("projects.json", "w", encoding="utf8") as f:

        json.dump(
            projects,
            f,
            indent=4,
            ensure_ascii=False
        )

    print(f"Generated {len(projects)} projects.")
    return projects